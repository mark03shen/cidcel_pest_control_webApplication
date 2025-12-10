# services/views.py
from django.shortcuts import render, get_object_or_404, redirect
# from django.contrib.auth.forms import AuthenticationForm 
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login, authenticate, get_backends
from django.views.decorators.http import require_POST
from urllib3 import request
from .forms import CustomUserCreationForm


from .models import Appointment
from django.http import JsonResponse
from django.core import serializers
from django.views.decorators.csrf import csrf_exempt

from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required, user_passes_test

from django.contrib.auth import logout

from .models import Product
from .forms import ProductForm
from django.db.models import Q 



def landing_page(request):
    feedbacks = Feedback.objects.all().order_by('-submitted_at')
    photos = UploadedPhoto.objects.all().order_by('-uploaded_at')
    return render(request, 'landing_page.html', {'feedbacks': feedbacks, 'photos': photos})


# this is for Admin SIGNUP and LOGIN
def is_admin(user):
    return user.is_staff

def admin_login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)

        if user is not None and user.is_staff:
            login(request, user)
            return redirect('admin_dashboard')
        else:
            messages.error(request, 'Invalid credentials or not an admin.')
    
    return render(request, 'admin_login.html')

@login_required
@user_passes_test(is_admin)
def admin_dashboard(request):
    return render(request, 'admin_pages/admin_dashboard.html')



User = get_user_model()

def admin_signup_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')

        # Check if the username already exists
        if User.objects.filter(username=username).exists():
            messages.error(request, 'Username already exists.')
        else:
            # Create the user with is_staff=True to give admin access
            user = User.objects.create_user(
                username=username,
                email=email,
                password=password
            )
            user.is_staff = True
            user.save()

            messages.success(request, 'Admin account created successfully. Please log in.')
            return redirect('admin_login')  # Replace with your actual login URL name

    return render(request, 'admin_signup.html')




# this is for customer signup
# Signup View (Auto-login after registration) 

def signup(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()

            # Set the backend manually
            backend = get_backends()[0]  # This assumes you want to use the first backend
            user.backend = f"{backend.__module__}.{backend.__class__.__name__}"

            login(request, user)
            return redirect('/signup/?success=1')  # triggers popup
    else:
        form = CustomUserCreationForm()
    return render(request, 'signup.html', {'form': form})


# Login View
def user_login(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(username=username, password=password)
        if user:
            login(request, user)
            return redirect('customer_dashboard')
    return render(request, 'login.html')


#this is a CUSTOMER DASHBOARD

def customer_dashboard(request):
    feedbacks = Feedback.objects.all().order_by('-submitted_at')
    photos = UploadedPhoto.objects.all().order_by('-uploaded_at')
    return render(request, 'customer/dashboard.html', {
        'user': request.user,
        'feedbacks': feedbacks,
        'photos': photos
    })

def dashboard(request):
    feedbacks = Feedback.objects.all().order_by('-submitted_at')
    return render(request, 'customer/dashboard.html', {'feedbacks': feedbacks})

def services(request):
    return render(request, 'customer/services.html')

def faqs(request):
    return render(request, 'customer/faqs.html')


#-- for customer feedback
from .models import Feedback
@login_required
def feedback(request):
    feedbacks = Feedback.objects.all().order_by('-submitted_at')  # newest first
    return render(request, 'customer/feedback.html', {'feedbacks': feedbacks})


def submit_feedback(request):
    if request.method == 'POST':
        category = request.POST.get('category')
        message = request.POST.get('feedback')
        rating = request.POST.get('rating')

        Feedback.objects.create(
            user=request.user if request.user.is_authenticated else None,
            category=category,
            message=message,
            rating=int(rating) if rating else 0
        )
        return redirect('feedback')

#--- for customer profile
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.forms import PasswordChangeForm
from .forms import ProfileUpdateForm, ProfilePictureForm

# @login_required
# def profile(request):
#     # Get appointments that are either Pending OR have a verification assignment
#     appointments = (
#         Appointment.objects.filter(
#             Q(status='Pending') | Q(verificationassignment__isnull=False)
#         )
#         .order_by( '-created_at')
#         .distinct()
#         .prefetch_related('verificationassignment_set')
#     )

#     return render(request, "customer/profile.html", {
#         "appointments": appointments
#     })
@login_required
def profile(request):
    # Only fetch appointments from the logged-in user
    appointments = (
        Appointment.objects.filter(
            user=request.user  # ← VERY IMPORTANT
        )
        .filter(
            Q(status='Pending') | Q(verificationassignment__isnull=False)
        )
        .order_by('-created_at')
        .distinct()
        .prefetch_related('verificationassignment_set')
    )

    return render(request, "customer/profile.html", {
        "appointments": appointments
    })


from .models import Receipt
@login_required
def upload_receipt(request, appointment_id):
    # Fetch the appointment by ID and only allow upload if payment method is Bank Transfer or GCash
    appointment = get_object_or_404(Appointment, id=appointment_id, payment_method__in=["Bank Transfer", "GCash"])

    if request.method == "POST" and request.FILES.get("receipt"):
        # Save the uploaded file directly to the appointment's receipt field
        appointment.receipt = request.FILES["receipt"]
        appointment.save()
        # return redirect("profile") 
        return redirect(f"{request.META.get('HTTP_REFERER', '/profile')}?uploaded=1")

    return redirect("profile")

@login_required
def delete_receipt(request, receipt_id):
    receipt = get_object_or_404(Receipt, id=receipt_id)
    
    if request.method == "POST":
        receipt.delete()
        messages.success(request, "Receipt deleted successfully.")
        return redirect(request.META.get('HTTP_REFERER', '/'))  # return to the page the user was on
    
    return redirect("/")


@login_required
def edit_profile(request):
    # ensure profile exists
    if not hasattr(request.user, 'profile'):
        from .models import Profile
        Profile.objects.create(user=request.user)

    if request.method == 'POST':
        u_form = ProfileUpdateForm(request.POST, instance=request.user)
        p_form = PasswordChangeForm(request.user, request.POST)
        pic_form = ProfilePictureForm(request.POST, request.FILES, instance=request.user.profile)

        if 'update_profile' in request.POST and u_form.is_valid():
            u_form.save()
            messages.success(request, 'Profile updated successfully!')
            return redirect('profile')

        if 'change_password' in request.POST and p_form.is_valid():
            user = p_form.save()
            update_session_auth_hash(request, user)
            messages.success(request, 'Password changed successfully!')
            return redirect('profile')

        if 'update_picture' in request.POST and pic_form.is_valid():
            pic_form.save()
            messages.success(request, 'Profile picture updated!')
            return redirect('profile')
    else:
        u_form = ProfileUpdateForm(instance=request.user)
        p_form = PasswordChangeForm(request.user)
        pic_form = ProfilePictureForm(instance=request.user.profile)

    return render(request, 'customer/edit_profile.html', {
        'u_form': u_form,
        'p_form': p_form,
        'pic_form': pic_form
    })



#----- for ADMIN HOME PAGE
import base64, mimetypes
@login_required
def admin_home(request):
    feedbacks = Feedback.objects.all().order_by('-submitted_at')[:10]
    notifications = TechnicianNotification.objects.select_related('technician', 'appointment').all().order_by('-timestamp')
    customer_notifications = CustomerNotification.objects.select_related("appointment").all().order_by('-timestamp')

    
    # fetch all technicians' latest photos (not just the logged-in user)
    photos = PhotoDocumentation.objects.select_related('technician').all().order_by('-created_at')[:20]
    photos_with_data = []
    for p in photos:
        try:
            with p.image.open('rb') as f:
                data = f.read()
            mime = mimetypes.guess_type(p.image.name)[0] or 'image/jpeg'
            p.data_uri = f"data:{mime};base64,{base64.b64encode(data).decode()}"
            photos_with_data.append(p)
        except Exception as e:
            print("Error reading image for base64:", e)
            
    # Payment receipts (only with Bank Transfer or GCash and uploaded receipt)
    receipts = Appointment.objects.filter(payment_method__in=["Bank Transfer", "GCash"], receipt__isnull=False).order_by('-date')

    receipts_with_data = []
    for r in receipts:
        try:
            with r.receipt.open('rb') as f:
                data = f.read()
            mime = mimetypes.guess_type(r.receipt.name)[0] or 'image/jpeg'
            r.data_uri = f"data:{mime};base64,{base64.b64encode(data).decode()}"
            receipts_with_data.append(r)
        except Exception as e:
            print("Error reading receipt image for base64:", e)
            
    
    return render(request, 'admin_pages/home.html', {
        "feedbacks": feedbacks,
        "notifications": notifications,
        "photos": photos,
        "receipts": receipts_with_data,
        "customer_notifications": customer_notifications, 
    })
    
    
    
@login_required
def admin_dashboard(request):
    feedbacks = Feedback.objects.all().order_by('-submitted_at')[:10]
    notifications = TechnicianNotification.objects.select_related('technician', 'appointment').all().order_by('-timestamp')
    customer_notifications = CustomerNotification.objects.select_related("appointment").all().order_by('-timestamp')
    
    # fetch all technicians' latest photos (not just the logged-in user)
    photos = PhotoDocumentation.objects.select_related('technician').all().order_by('-created_at')[:20]
    photos_with_data = []
    for p in photos:
        try:
            with p.image.open('rb') as f:
                data = f.read()
            mime = mimetypes.guess_type(p.image.name)[0] or 'image/jpeg'
            p.data_uri = f"data:{mime};base64,{base64.b64encode(data).decode()}"
            photos_with_data.append(p)
        except Exception as e:
            print("Error reading image for base64:", e)
            
    # Payment receipts (only with Bank Transfer or GCash and uploaded receipt)
    receipts = Appointment.objects.filter(payment_method__in=["Bank Transfer", "GCash"], receipt__isnull=False).order_by('-date')

    receipts_with_data = []
    for r in receipts:
        try:
            with r.receipt.open('rb') as f:
                data = f.read()
            mime = mimetypes.guess_type(r.receipt.name)[0] or 'image/jpeg'
            r.data_uri = f"data:{mime};base64,{base64.b64encode(data).decode()}"
            receipts_with_data.append(r)
        except Exception as e:
            print("Error reading receipt image for base64:", e)
            
    
    return render(request, 'admin_pages/home.html', {
        "feedbacks": feedbacks,
        "notifications": notifications,
        "photos": photos,
        "receipts": receipts_with_data,
        "customer_notifications": customer_notifications, 
    })
    
    
from .models import ChatMessage

@login_required
def send_message(request):
    if request.method == "POST":
        msg = request.POST.get("message")
        if msg.strip():
            ChatMessage.objects.create(sender=request.user, message=msg, is_admin=False)
        return JsonResponse({"status": "ok"})

@login_required
def load_messages(request):
    messages = ChatMessage.objects.all()
    data = [
        {
            "sender": "Admin" if m.is_admin else m.sender.username,
            "message": m.message,
            "timestamp": m.timestamp.strftime("%Y-%m-%d %H:%M"),
            "is_admin": m.is_admin
        }
        for m in messages
    ]
    return JsonResponse({"messages": data})


@login_required
def admin_send_message(request):
    if request.method == "POST":
        msg = request.POST.get("message")
        if msg.strip():
            ChatMessage.objects.create(sender=request.user, message=msg, is_admin=True)
        return JsonResponse({"status": "ok"})


from .forms import UploadPhotoForm
from .models import UploadedPhoto

@login_required
def upload_photo(request):
    if request.method == "POST":
        form = UploadPhotoForm(request.POST, request.FILES)
        if form.is_valid():
            upload = form.save(commit=False)
            upload.uploaded_by = request.user
            upload.save()
            return redirect('upload')  # reload page after upload
    else:
        form = UploadPhotoForm()

    photos = UploadedPhoto.objects.all().order_by('-uploaded_at')
    return render(request, "admin_pages/upload.html", {"form": form, "photos": photos})

@login_required
def delete_picture(request, photo_id):
    photo = get_object_or_404(UploadedPhoto, id=photo_id)
    if request.user.is_superuser or request.user == photo.uploaded_by:
        photo.delete()
    return redirect('upload')


#----- this is for BOOK APPOINTMENT BUTTON

@login_required
def book_appointment(request):
    if request.method == "POST":
        service = request.POST.get("service")
        date = request.POST.get("date")
        time = request.POST.get("time")

        # Save appointment for logged-in user
        appointment = Appointment.objects.create(
            service=service,
            date=date,
            time=time,
            email=request.user.email,
            status="Pending"
        )

        # Store confirmation details in session
        request.session["appointment_message"] = {
            "name": request.user.username,
            "service": service,
            "date": date,
            "time": time,
        }

        return redirect("profile")  # Go to profile after booking

    return render(request, "admin_pages/book_appointment.html")


from decimal import Decimal
from .models import CustomerNotification

@login_required
def create_appointment(request):
    if request.method == 'POST':
        # prefer logged-in user's data where appropriate
        client_name = request.POST.get('client_name') or request.user.get_full_name() or request.user.username
        client_address = request.POST.get('client_address', '')
        email = request.POST.get('email') or request.user.email
        mobile = request.POST.get('mobile', '')
        
        # ✅ clean estimated_price (remove ₱ and commas)
        raw_price = request.POST.get("estimated_price", "")
        cleaned_price = None
        if raw_price:
            try:
                cleaned_price = Decimal(raw_price.replace("₱", "").replace(",", "").strip())
            except:
                cleaned_price = None

        appointment = Appointment.objects.create(
            user=request.user, 
            client_name=client_name,
            client_address=client_address,
            land_use_category=request.POST.get('land_use_category', ''),
            email = request.POST.get('email') or request.user.email,
            mobile=mobile,
            job_location=request.POST.get('job_location', ''),
            municipality=request.POST.get('municipality', ''),
            barangay=request.POST.get('barangay', ''),
            street=request.POST.get('street', ''),
            house_number=request.POST.get('house_number', ''),
            total_area=request.POST.get('total_area') or 0,
            service=request.POST.get('service', ''),
            date=request.POST.get('date'),
            time=request.POST.get('time'),
            contract_period=request.POST.get('contract_period', ''),
            payment_method=request.POST.get('payment_method', ''),
            estimated_price=cleaned_price, 
            status='Pending'
        )
        appointment.save()
        
        # ✅ CREATE NOTIFICATION FOR ADMIN
        CustomerNotification.objects.create(
            appointment=appointment,
            client_name=appointment.client_name,
            message=f"New appointment request from {client_name}"
        )

        
        # store confirmation in session so profile view can show it once
        request.session['appointment_message'] = {
            'name': client_name,
            'service': appointment.service,
            'date': str(appointment.date),
            'time': str(appointment.time),
        }

        # Redirect to profile so user can immediately see the confirmation
        return redirect('profile')

    # If not POST, redirect back to the booking page
    return redirect('book_appointment')

def cancel_booking(request):
    # Optional: clear session or temporary data here if needed
    return redirect('customer_dashboard') 


@login_required
def recent_appointment(request):
    # Get the most recent appointment for the logged-in user
    appointment = Appointment.objects.filter(user=request.user).order_by('-created_at').first()
    
    if appointment:
        data = {
            "client_name": appointment.client_name,
            "client_address": appointment.client_address,
            "land_use_category": appointment.land_use_category,
            "email": appointment.email,
            "mobile": appointment.mobile,
            "job_location": appointment.job_location,
            "municipality": appointment.municipality,
            "barangay": appointment.barangay,
            "street": appointment.street,
            "house_number": appointment.house_number,
            "total_area": float(appointment.total_area or 0),
            "service": appointment.service,
            "contract_period": appointment.contract_period,
            "estimated_price": float(appointment.estimated_price or 0),
            "date": appointment.date.strftime("%Y-%m-%d"),
            "time": appointment.time.strftime("%H:%M"),
            "payment_method": appointment.payment_method,
        }
        return JsonResponse({"success": True, "appointment": data})
    else:
        return JsonResponse({"success": False, "message": "No recent appointment found"})


# @login_required
# def analytics(request):
#     notifications = TechnicianNotification.objects.select_related('technician', 'appointment').all().order_by('-timestamp')
#     return render(request, 'admin_pages/analytics.html', {
#         "notifications": notifications
#     }) 
from django.db.models.functions import TruncMonth, ExtractYear
from django.db.models import Count, Sum
import datetime

@login_required
def analytics(request):
    notifications = TechnicianNotification.objects.select_related('technician', 'appointment').all().order_by('-timestamp')
    customer_notifications = CustomerNotification.objects.select_related("appointment").all().order_by('-timestamp')
    completed_status = "Completed"  # or "Done" depending on your DB

    # 1. Service Demand Forecast
    service_demand = (
        Appointment.objects.filter(status=completed_status)
        .values("service")
        .annotate(
            total=Count("id"),
            revenue=Sum("estimated_price")
        )
        .order_by("-total")
    )
    
    total_jobs = Appointment.objects.filter(status=completed_status).count()
    total_revenue = (
        Appointment.objects.filter(status=completed_status)
        .aggregate(total=Sum("estimated_price"))["total"] or 0
    )

    # 2. Completed Jobs Over Time (last 12 months)
    today = datetime.date.today()
    twelve_months_ago = today - datetime.timedelta(days=365)

    jobs_over_time = (
        Appointment.objects.filter(status=completed_status, date__gte=twelve_months_ago)
        .annotate(month=TruncMonth("date"))
        .annotate(year=ExtractYear("date"))
        .values("month", "year")
        .annotate(total=Count("id"))
        .order_by("year", "month")
    )
    

    # 3. Land Categories (distribution of booked appointments)
    land_data = (
        Appointment.objects.filter(status=completed_status)
        .values("land_use_category")
        .annotate(total=Count("id"))
        .order_by("-total")
    )

    total_land_count = sum(item["total"] for item in land_data) or 1  # avoid divide by zero

    land_categories = [
        {
            "land_category": item["land_use_category"],
            "percentage": round((item["total"] / total_land_count) * 100, 2),
        }
        for item in land_data
    ]

    context = {
        "notifications": notifications,
        "total_revenue": total_revenue,
        "service_demand": service_demand,
        "total_jobs": total_jobs,
        "jobs_over_time": jobs_over_time,
        "land_categories": land_categories,
        "customer_notifications": customer_notifications, 
    }
    return render(request, "admin_pages/analytics.html", context)



@login_required
def pest_activity(request):
    notifications = TechnicianNotification.objects.select_related('technician', 'appointment').all().order_by('-timestamp')
    customer_notifications = CustomerNotification.objects.select_related("appointment").all().order_by('-timestamp')
    
    return render(request, 'admin_pages/pest_activity.html', {
        "notifications": notifications,
        "customer_notifications": customer_notifications, 
    }) 

#--- this is for CREATING HEAT MAP
# services/views.py
import os, re
import pandas as pd
from django.conf import settings
from django.shortcuts import render
from django.http import JsonResponse

def parse_coord(s):
    """Parses strings like '13°9421′N' -> 13.9421
       Fallbacks to extracting a float if format differs."""
    if pd.isna(s):
        return None
    s = str(s).strip()
    sign = -1 if any(ch in s for ch in ('S','s','W','w')) else 1
    m = re.match(r'^\s*(\d+)[°\s]*([0-9]+)', s)
    if m:
        deg = int(m.group(1))
        frac = m.group(2)
        try:
            return sign * (deg + int(frac) / (10 ** len(frac)))
        except Exception:
            return None
    m2 = re.search(r'(-?\d+\.\d+|-?\d+)', s)
    if m2:
        return float(m2.group(0))
    return None

def pest_activity_view(request):
    # Renders the page which will hold the Leaflet map
    notifications = TechnicianNotification.objects.select_related('technician', 'appointment').all().order_by('-timestamp')
    customer_notifications = CustomerNotification.objects.select_related("appointment").all().order_by('-timestamp')
    return render(request, 'admin_pages/pest_activity.html', {
        "notifications": notifications,
        "customer_notifications": customer_notifications,
    })

def pest_activity_data(request):
    # Returns JSON with points [[lat, lon], ...]
    file_path = os.path.join(settings.BASE_DIR, 'services', 'data',
                             'SERVICE REPORT CIDSEL PEST CONTROL SERVICES.xlsx')
    df = pd.read_excel(file_path, sheet_name='Sariaya', engine='openpyxl')
    df['lat_dec'] = df['LATITUDE'].apply(parse_coord)
    df['lon_dec'] = df['LONGITUDE'].apply(parse_coord)
    pts = df.dropna(subset=['lat_dec', 'lon_dec'])[['lat_dec', 'lon_dec']].values.tolist()
    return JsonResponse({'points': pts})

#--- this is for CREATING HEAT MAP




#-----------for APPOINTMENT-------------#
@login_required
def appointment(request):
    technicians = CustomUser.objects.filter(is_technician=True)
    
    tech_data = []
    for tech in technicians:
        status, _ = TechnicianStatus.objects.get_or_create(technician=tech)
        tech_data.append({
            "id": tech.id,
            "name": tech.username,
            "online": status.is_online()
        })
        
    appointments = Appointment.objects.filter(
        Q(status='Pending') | Q(verificationassignment__isnull=False)
    ).order_by('-created_at').distinct().prefetch_related('verificationassignment_set')


    notifications = TechnicianNotification.objects.all().order_by('-timestamp')
    customer_notifications = CustomerNotification.objects.select_related("appointment").all().order_by('-timestamp')
                            
    return render(request, 'admin_pages/appointment.html', {
        'appointments': appointments,
        'notifications': notifications,
        'customer_notifications': customer_notifications,
        "technicians": tech_data,
    })


from django.db.models import Case, When, IntegerField

@login_required
def admin_appointment_view(request):
    technicians = CustomUser.objects.filter(is_technician=True)
    
    tech_data = []
    for tech in technicians:
        status, _ = TechnicianStatus.objects.get_or_create(technician=tech)
        tech_data.append({
            "id": tech.id,
            "name": tech.username,
            "online": status.is_online()
        })
    
    appointments = Appointment.objects.order_by(
        Case(
            When(status="Pending", then=0),
            default=1,  
            output_field=IntegerField(),
        ),
        '-created_at'
    )
    return render(request, 'admin_pages/appointment.html', {
        'appointments': appointments,
        "technicians": tech_data,
    })


# add near other small JSON endpoints in services/views.py
from django.contrib.auth.decorators import login_required
@login_required
def get_progress_steps(request):
    """
    AJAX endpoint: given appointment_id (GET),
    returns JSON: { success: True, steps: [...], current_status: "..." }
    """
    appointment_id = request.GET.get("appointment_id")
    if not appointment_id:
        return JsonResponse({"success": False, "error": "No appointment_id provided"})

    try:
        # try to find the latest assignment for the appointment
        assignment = VerificationAssignment.objects.filter(appointment_id=appointment_id).order_by('-assigned_at').first()
        if not assignment:
            return JsonResponse({"success": False, "error": "No assignment found for this appointment"})

        service = assignment.appointment.service or ""
        period = (assignment.appointment.contract_period or "").lower()

        # Build step list exactly like verification_route.html
        if service == "General Pest Control":
            steps = ["Initial Visit", "Follow Up Visit", "Last Visit", "Completed"]

        elif service == "Termite Spot Treatment":
            steps = ["Initial Visit", "Follow Up Visit", "Last Visit", "Completed"]

        elif service == "Termite Comprehensive":
            if "1" in period:
                steps = ["Initial Visit", "2nd Follow Up Visit", "Last Visit", "Completed"]
            else:
                steps = ["Initial Visit", "2nd Follow Up Visit", "3rd Follow Up Visit",
                         "4th Follow Up Visit", "5th Follow Up Visit", "Last Visit", "Completed"]
        else:
            # fallback generic steps
            steps = ["Initial Visit", "Follow Up Visit", "Last Visit", "Completed"]

        return JsonResponse({
            "success": True,
            "steps": steps,
            "current_status": assignment.progress_status 
        })

    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})

#--- for uploaded progress photo
def get_progress_photos(request):
    appointment_id = request.GET.get("appointment_id")

    if not appointment_id:
        return JsonResponse({"success": False, "error": "No appointment id provided."})

    try:
        # Find all assignments linked to this appointment
        assignments = VerificationAssignment.objects.filter(appointment_id=appointment_id)

        # Fetch progress entries
        progress_items = TechnicianProgress.objects.filter(
            assignment__in=assignments
        ).order_by('-created_at')

        photos = []
        for item in progress_items:
            photos.append({
                "image_url": item.photo.url,
                "timestamp": item.created_at.strftime("%b %d, %Y • %I:%M %p"),
                "status_label": item.progress_status,  # Initial Visit, etc.
                "technician": item.assignment.technician.username  # Who uploaded it
            })

        return JsonResponse({"success": True, "photos": photos})

    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})




@csrf_exempt
@login_required
@require_POST
def update_appointment_status(request, appointment_id):
    try:
        appointment = Appointment.objects.get(id=appointment_id)
        data = json.loads(request.body.decode("utf-8"))

        status = data.get("status")
        service = data.get("service")
        contract_period = data.get("contract_period")

        if not status:
            return JsonResponse({"success": False, "error": "No status provided"})

        # ✅ Update appointment status with the dropdown selection
        appointment.status = status
        appointment.save()

        return JsonResponse({
            "success": True,
            "new_status": status,
            "service": service,
            "contract_period": contract_period,
        })

    except Appointment.DoesNotExist:
        return JsonResponse({"success": False, "error": "Appointment not found"})
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})

#-----------for admin APPOINTMENT-------------#


@login_required
def customers(request):
    notifications = TechnicianNotification.objects.select_related('technician', 'appointment').all().order_by('-timestamp')
    customer_notifications = CustomerNotification.objects.select_related("appointment").all().order_by('-timestamp')
    
    # ✅ Fetch most recent 10 completed appointments
    recent_customers = (
        Appointment.objects.filter(status="Completed")
        .order_by("-date", "-time")[:10]
    )

    return render(request, "admin_pages/customers.html", {
        "notifications": notifications,
        "recent_customers": recent_customers,
        "customer_notifications": customer_notifications, 
    })


@login_required
def customers_review(request):
    notifications = TechnicianNotification.objects.select_related('technician', 'appointment').all().order_by('-timestamp')
    customer_notifications = CustomerNotification.objects.select_related("appointment").all().order_by('-timestamp')
    
    # ✅ fetch completed appointments
    customers = (
        Appointment.objects.filter(status="Completed")
        .order_by("-date", "-time")
    )

    return render(request, "admin_pages/customers_review.html", {
        "notifications": notifications,
        "customers": customers,
        "customer_notifications": customer_notifications, 
    })


@login_required
def get_customer_detail(request, pk):
    try:
        cust = Appointment.objects.get(pk=pk, status__in=["Booked", "Completed"])
        data = {
            "client_name": cust.client_name,
            "service": cust.service,
            "estimated_price": cust.estimated_price,
            "date": cust.date.strftime("%B %d, %Y"),
            "time": cust.time.strftime("%H:%M"),
            "status": cust.status,   # ✅ include status
            "email": cust.email,
            "mobile": cust.mobile,
        }
        return JsonResponse(data)
    except Appointment.DoesNotExist:
        return JsonResponse({"error": "Customer not found"}, status=404)




#-----------for INVENTORY-------------#
@login_required
def inventory(request):
    if request.method == 'POST':
        form = ProductForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            # defensive server-side enforcement
            if obj.category == "Materials":
                obj.measurement = "L"
            else:
                obj.measurement = None
            obj.save()
            return redirect('inventory')
    else:
        form = ProductForm()

    products = Product.objects.all()
    # Auto-assign stock status (computed, not stored)
    for p in products:
        if p.quantity <= 0:
            p.status = "Out of Stock"
        elif p.quantity <= 5:
            p.status = "Low Stock"
        else:
            p.status = "Available"
    notifications = TechnicianNotification.objects.select_related('technician', 'appointment').all().order_by('-timestamp')
    customer_notifications = CustomerNotification.objects.select_related("appointment").all().order_by('-timestamp')

    return render(request, 'admin_pages/inventory.html', {
        'form': form,
        'products': products,
        "notifications": notifications,
        "customer_notifications": customer_notifications,
    }) 

def inventory_list(request):
    notifications = TechnicianNotification.objects.select_related('technician', 'appointment').all().order_by('-timestamp')
    customer_notifications = CustomerNotification.objects.select_related("appointment").all().order_by('-timestamp')
    products = Product.objects.all()
    form = ProductForm()
    if request.method == "POST":
        form = ProductForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('inventory')
    return render(request, 'admin_pages/inventory.html', {
        'products': products, 
        'form': form, 
        "notifications": notifications, 
        "customer_notifications": customer_notifications,
    })

def product_update(request, pk):
    product = get_object_or_404(Product, pk=pk)
    form = ProductForm(request.POST or None, instance=product)
    if form.is_valid():
        obj = form.save(commit=False)
        if obj.category == "Materials":
            obj.measurement = "L"
        else:
            obj.measurement = None
        obj.save()
        return redirect('inventory')
    return render(request, 'inventory_edit.html', {'form': form, 'product': product})


def product_delete(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        product.delete()
        return redirect('inventory')
    return render(request, 'inventory_delete.html', {'product': product})


def update_inventory_on_assign():
    from .models import Product

    # Decrease all materials by 1 (if quantity > 0)
    materials = Product.objects.filter(category="Materials")
    for item in materials:
        if item.quantity > 0:
            item.quantity -= 1
            item.save()

    # Mark all equipment as used
    equipment = Product.objects.filter(category="Equipment")
    equipment.update(is_used=True)


def get_recommended_items(request):
    assignment_id = request.GET.get("assignment_id")

    try:
        assignment = VerificationAssignment.objects.get(id=assignment_id)
        service = assignment.appointment.service

        # -----------------------
        #   PRESET LISTS HERE
        # -----------------------

        if service == "General Pest Control":
            materials = [
                "Agenda 1L (optional for GPC)",
                "Permiotor 1L",
                "Diptor Doo 1L",
                "Resigene 1L",
                "ISOPTEX 1L (optional)"
            ]
            equipment = [
                "Power Sprayer", "Hand Sprayer", "Mistng Machine", "Fogging Machine",
                "Drum/Container", "Hose", "Nozzle", "Filter cloth", "Flashlight",
                "Respirator", "Rubber Gloves", "Cotton Gloves", "Eye Goggles",
                "Hard Hat", "Rubber Boots", "Safety Shoes"
            ]

        elif service == "Termite Spot Treatment":
            materials = [
                "Agenda 1L",
                "Permiotor 1L",
                "ISOPTEX 1L (if using fipronil-based spot treatment)"
            ]
            equipment = [
                "Hand Sprayer", "Injector", "Drill Machine", "Power Sprayer (optional)",
                "Drum/Container", "Hose", "Nozzle", "Flashlight", "Respirator",
                "Rubber Gloves", "Eye Goggles", "Cotton Gloves",
                "Safety Shoes", "Hard Hat"
            ]

        elif service == "Termite Comprehensive":
            materials = [
                "ISOPTEX 1L",
                "Agenda 1L",
                "Permiotor 1L"
            ]
            equipment = [
                "Power Sprayer", "Injector", "Drill Machine", "PVC Pipes",
                "Filter Cloth", "Drum/Container", "Hose", "Nozzle", "Flashlight",
                "Mistng Machine (optional)", "Respirator", "Rubber Gloves",
                "Eye Goggles", "Hard Hat", "Rubber Boots", "Safety Shoes"
            ]

        else:
            materials = []
            equipment = []

        return JsonResponse({
            "success": True,
            "materials": materials,
            "equipment": equipment
        })

    except VerificationAssignment.DoesNotExist:
        return JsonResponse({"success": False, "error": "Assignment not found"})
    
#-----------for INVENTORY-------------#


#-----------for SALES-------------#
from django.db.models import Sum, Count

@login_required
def sales(request):
    notifications = TechnicianNotification.objects.select_related('technician', 'appointment').all().order_by('-timestamp')
    customer_notifications = CustomerNotification.objects.select_related("appointment").all().order_by('-timestamp')
    
    # ✅ Filter only completed appointments
    completed_appointments = Appointment.objects.filter(status="Completed")

    total_revenue = completed_appointments.aggregate(Sum("estimated_price"))["estimated_price__sum"] or 0
    total_bookings = completed_appointments.count()
    avg_sale_value = (total_revenue / total_bookings) if total_bookings > 0 else 0

    # ✅ Sales breakdown by service
    general_sales = completed_appointments.filter(service="General Pest Control").aggregate(Sum("estimated_price"))["estimated_price__sum"] or 0
    termite_comprehensive_sales = completed_appointments.filter(service="Termite Comprehensive").aggregate(Sum("estimated_price"))["estimated_price__sum"] or 0
    termite_spot_sales = completed_appointments.filter(service="Termite Spot Treatment").aggregate(Sum("estimated_price"))["estimated_price__sum"] or 0

    # ✅ Top service by revenue
    service_revenue_map = {
        "General Pest Control": general_sales,
        "Termite Comprehensive": termite_comprehensive_sales,
        "Termite Spot Treatment": termite_spot_sales,
    }
    top_service = max(service_revenue_map, key=service_revenue_map.get) if total_revenue > 0 else "N/A"

    return render(request, "admin_pages/sales.html", {
        "notifications": notifications,
        "total_revenue": total_revenue,
        "total_bookings": total_bookings,
        "avg_sale_value": avg_sale_value,
        "general_sales": general_sales,
        "termite_comprehensive_sales": termite_comprehensive_sales,
        "termite_spot_sales": termite_spot_sales,
        "top_service": top_service,
        "completed_appointments": completed_appointments,
        "customer_notifications": customer_notifications,  
    })

#---- this is for REPORTS -----
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
 
@login_required
def reports(request):
    notifications = TechnicianNotification.objects.select_related('technician', 'appointment').all().order_by('-timestamp')
    customers = Appointment.objects.filter(status="Completed").order_by("-date", "-time")
    customer_notifications = CustomerNotification.objects.select_related("appointment").all().order_by('-timestamp')
    
    # Example sales data
    sales_data = (
        Appointment.objects.filter(status='Completed')
        .values('service')
        .annotate(total_jobs=Count('id'), total_revenue=Sum('estimated_price'))
        .order_by('service')
    )

    # Calculate totals
    total_jobs = sum(item['total_jobs'] for item in sales_data)
    total_revenue = sum(item['total_revenue'] or 0 for item in sales_data)

    context = {
        "notifications": notifications,
        "customers": customers,
        "sales_data": sales_data,
        "total_jobs": total_jobs,
        "total_revenue": total_revenue,
        "customer_notifications": customer_notifications, 
    }

    return render(request, 'admin_pages/reports.html', context)

def sales_report(request):
    monthly_sales = (
        Appointment.objects.filter(status="Completed")
        .annotate(month=TruncMonth("created_at"))
        .values("month")
        .annotate(total=Sum("price"))
        .order_by("-month")
    )

    # Format months nicely
    for item in monthly_sales:
        item["month_name"] = item["month"].strftime("%B %Y")

    return render(request, "reports.html", {
        "monthly_sales": monthly_sales
    })
    
def sales_by_month(request):
    month = request.GET.get("month")  # format: yyyy-mm

    appointments = Appointment.objects.filter(
        status="Completed",
        created_at__year=month[:4],
        created_at__month=month[5:7]
    ).values("customer_name", "service_type", "created_at", "price")

    return JsonResponse(list(appointments), safe=False)

# Export Sales to Excel
@login_required
def export_sales_report(request):
    completed_status = "Completed"
    sales_data = (
        Appointment.objects.filter(status=completed_status)
        .values('service')
        .annotate(total_jobs=Count('id'), total_revenue=Sum('estimated_price'))
        .order_by('-total_jobs')
    )

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    # Header
    headers = ['Service', 'Total Jobs', 'Revenue (₱)']
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    # Fill data
    for row_num, item in enumerate(sales_data, 2):
        ws.cell(row=row_num, column=1, value=item['service'])
        ws.cell(row=row_num, column=2, value=item['total_jobs'])
        ws.cell(row=row_num, column=3, value=float(item['total_revenue'] or 0))

    # Total row
    total_jobs = sum(item['total_jobs'] for item in sales_data)
    total_revenue = sum(float(item['total_revenue'] or 0) for item in sales_data)
    ws.cell(row=len(sales_data)+2, column=1, value="Total")
    ws.cell(row=len(sales_data)+2, column=2, value=total_jobs)
    ws.cell(row=len(sales_data)+2, column=3, value=total_revenue)

    # Adjust column width
    for col in range(1, 4):
        ws.column_dimensions[get_column_letter(col)].width = 20

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=Sales_Report.xlsx'
    wb.save(response)
    return response



def logout_view(request):
    logout(request)
    return redirect('login')  # Adjust to your login page route




#------- this is for technician signup and login -------
from .forms import TechnicianSignUpForm
from .models import CustomUser


def technician_login(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)

        if user is not None and user.is_technician:
            login(request, user)
            messages.success(request, f"Welcome {user.username}!")
            return redirect('home_page')  # Update this to your actual dashboard/view
        else:
            messages.error(request, 'Invalid credentials or not authorized as technician.')

    return render(request, 'technician_login.html')


def technician_signup(request):
    if request.method == 'POST':
        form = TechnicianSignUpForm(request.POST)
        if form.is_valid():
            user = form.save()
            messages.success(request, "Account created successfully. Please log in.")
            return redirect('technician_login')
        else:
            print(form.errors)  # 👈 Add this to see actual errors in terminal
            messages.error(request, "Please correct the errors below.")
    else:
        form = TechnicianSignUpForm()

    return render(request, 'technician_signup.html', {'form': form})


from services.models import TechnicianStatus, timezone
@login_required
def technician_dashboard(request):
    if request.user.is_authenticated and request.user.is_technician:
        TechnicianStatus.objects.update_or_create(
            technician=request.user,
            defaults={'last_active': timezone.now()}
        )
        
    if not request.user.is_technician:
        return redirect('technician_login')  # restrict access
    return render(request, 'technician_pages/technician_dashboard.html')



#------ for technician content ------
@login_required
def home_page(request):
    notifications = TechnicianNotification.objects.filter(technician=request.user, is_read=False).order_by('-timestamp')[:5]
    return render(request, 'technician_pages/home_page.html', {'notifications': notifications})

@login_required
def update_technician_status(request):
    tech = request.user
    if not tech.is_technician:
        return JsonResponse({"success": False})

    status, created = TechnicianStatus.objects.get_or_create(technician=tech)
    status.save()  # auto-updates last_active
    return JsonResponse({"success": True})

#---------- for NOTIFICATIONS FOR TECHNICIAN
@login_required
def notifications(request):
    all_notifications = TechnicianNotification.objects.filter(technician=request.user).order_by('-timestamp')
    return render(request, 'technician_pages/notifications.html', {
        'notifications': all_notifications
    })

@login_required
def mark_notification_read(request, notif_id):
    notification = get_object_or_404(TechnicianNotification, id=notif_id, technician=request.user)
    notification.is_read = True
    notification.save()
    return redirect('notifications')  # go to notifications.html


# @login_required
# def technician_notifications(request):
#     notif_id = request.GET.get('notif_id')
#     if notif_id:
#         try:
#             note = TechnicianNotification.objects.get(id=notif_id, technician=request.user)
#             note.is_read = True
#             note.save()
#             appointment = note.appointment
#         except TechnicianNotification.DoesNotExist:
#             note = None
#             appointment = None
#     else:
#         note = None
#         appointment = None

#     return render(request, 'technician_pages/notifications.html', {
#         'notification': note,
#         'appointment': appointment
#     })
    
    
    
#---------- for PHOTO DOCUMENTATION
from .forms import PhotoDocumentationForm
from services.models import PhotoDocumentation

@login_required
def photo_documentation(request):
    if request.method == 'POST':
        form = PhotoDocumentationForm(request.POST, request.FILES)
        if form.is_valid():
            photo = form.save(commit=False)
            photo.technician = request.user  # link photo to logged-in technician
            photo.save()

            # ✅ Debugging message
            messages.success(request, f"Photo saved! ID={photo.id}, User={photo.technician.username}")

            print("DEBUG >>> Saved Photo:", photo.id, photo.technician, photo.image.url)

            return redirect('photo_documentation')  # refresh page after upload
        else:
            # show form errors in console
            print("DEBUG >>> Form errors:", form.errors)
            messages.error(request, "Upload failed. Please check the form.")
    else:
        form = PhotoDocumentationForm()

    # fetch technician’s own photos
    photos = PhotoDocumentation.objects.filter(technician=request.user).order_by('-created_at')

    return render(request, 'technician_pages/photo_documentation.html', {
        'form': form,
        'photos': photos
    })

#--- delete photo
@login_required
def delete_photo(request, photo_id):
    photo = get_object_or_404(PhotoDocumentation, id=photo_id, technician=request.user)
    if request.method == "POST":
        photo.delete()
        messages.success(request, "Photo removed successfully.")
    return redirect('photo_documentation')


from .models import VerificationAssignment
from .models import CustomUser
from .models import TechnicianNotification
import json


#--- for verification route/Job assingment
@login_required
@user_passes_test(lambda u: u.is_technician)
def verification_route(request):
    highlight_id = request.GET.get('highlight_id')
    assignments = VerificationAssignment.objects.select_related('technician', 'appointment') \
        .filter(technician=request.user).order_by('-assigned_at')

    return render(request, 'technician_pages/verification_route.html', {
        'assignments': assignments,
        'highlight_id': highlight_id
    })


#-- for Notification to admin
@csrf_exempt
@login_required
@user_passes_test(lambda u: u.is_staff)
def assign_verification_ajax(request):
    if request.method == 'POST':
        appointment_id = request.POST.get('appointment_id')
        technician_id = request.POST.get('technician_id')
        

        try:
            appointment = Appointment.objects.get(id=appointment_id)
            technician = CustomUser.objects.get(id=technician_id, is_technician=True)

            obj, created = VerificationAssignment.objects.get_or_create(
                technician=technician,
                appointment=appointment
            )

            if created:
                update_inventory_on_assign()
                # Only notify that a job was assigned
                TechnicianNotification.objects.create(
                    technician=technician,
                    appointment=appointment,
                    message=f"You have been assigned a new job: {appointment.client_name} - {appointment.service}"
                )

            return JsonResponse({'success': True, 'status': 'assigned' if created else 'already_assigned'})
        
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

    return JsonResponse({'success': False, 'error': 'Invalid method'})

#--- for accepting the job
@csrf_exempt
@login_required
def technician_accept_job(request):
    if request.method == "POST":
        assignment_id = request.POST.get("assignment_id")
        try:
            assignment = VerificationAssignment.objects.get(id=assignment_id, technician=request.user)
            assignment.approval_status = "approved"   # ✅ only approval_status
            assignment.save()

            TechnicianNotification.objects.create(
                technician=request.user,
                appointment=assignment.appointment,
                message=f"The technician accepted the job request for {assignment.appointment.client_name} - {assignment.appointment.service}"
            )

            return JsonResponse({"success": True, "status": "approved"})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})
    return JsonResponse({"success": False, "error": "Invalid method"})


#--- for updating progress (dropdown)
def update_verification_status(request):
    if request.method == "POST":
        assignment_id = request.POST.get("assignment_id")
        status = request.POST.get("status")

        # ✅ UPDATE INVENTORY HERE
        update_inventory_on_assign()
        
        try:
            assignment = VerificationAssignment.objects.get(id=assignment_id)
            assignment.progress_status = status   # ✅ use progress_status, not approval_status
            assignment.save()

            # ✅ also sync to appointment if needed
            appointment = assignment.appointment
            appointment.status = status
            appointment.save()

            return JsonResponse({"success": True})
        except VerificationAssignment.DoesNotExist:
            return JsonResponse({"success": False, "error": "Assignment not found"})
    return JsonResponse({"success": False, "error": "Invalid request"})


from datetime import timedelta
# @login_required 
# @user_passes_test(lambda u: u.is_staff)
# def get_technicians(request):
#     technicians = CustomUser.objects.filter(is_technician=True).values('id', 'username')
#     return JsonResponse({'technicians': list(technicians)})
@login_required 
@user_passes_test(lambda u: u.is_staff)
def get_technicians(request):
    technicians = CustomUser.objects.filter(is_technician=True)
    tech_list = []

    for tech in technicians:
        status, _ = TechnicianStatus.objects.get_or_create(technician=tech)
        tech_list.append({
            'id': tech.id,
            'username': tech.username,
            'online': (timezone.now() - status.last_active) < timedelta(minutes=2)
        })

    return JsonResponse({'technicians': tech_list})


from .models import TechnicianProgress
@require_POST
def upload_progress_photo(request):
    try:
        assignment_id = request.POST.get("assignment_id")
        progress_status = request.POST.get("progress_status")
        description = request.POST.get("description")
        photo = request.FILES.get("photo")

        if not assignment_id or not progress_status or not description or not photo:
            return JsonResponse({"success": False, "error": "Missing fields."})

        assignment = VerificationAssignment.objects.get(id=assignment_id)

        # Save uploaded progress
        TechnicianProgress.objects.create(
            assignment=assignment,
            progress_status=progress_status,
            description=description,
            photo=photo
        )

        # Update the assignment progress_status too
        assignment.progress_status = progress_status
        assignment.save()

        return JsonResponse({"success": True})

    except VerificationAssignment.DoesNotExist:
        return JsonResponse({"success": False, "error": "Assignment not found."})

    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})



# fordissmising notif in technician home_page
@csrf_exempt
def dismiss_notification(request):
    if request.method == 'POST':
        notif_id = request.POST.get('notif_id')
        try:
            notif = TechnicianNotification.objects.get(id=notif_id)
            notif.delete()
            return JsonResponse({'status': 'success'})
        except TechnicianNotification.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Notification not found'})
    return JsonResponse({'status': 'error', 'message': 'Invalid request'})


#--- for single notification modal in admin notif
@login_required
def get_notification_detail(request):
    notif_id = request.GET.get('id')
    try:
        note = TechnicianNotification.objects.select_related('appointment', 'technician').get(id=notif_id)
        appointment = note.appointment
        data = {
            'technician': note.technician.username,
            'message': note.message,
            'timestamp': format(note.timestamp, "M d, Y H:i"),
        }

        if appointment:
            data.update({
                'appointment_id': appointment.id,  # ✅ Add this line
                'client_name': appointment.client_name,
                'service': appointment.service,
                'date': str(appointment.date),
                'time': str(appointment.time),
                'email': appointment.email,
                'mobile': appointment.mobile,
            })

        return JsonResponse(data)
    except TechnicianNotification.DoesNotExist:
        return JsonResponse({'error': 'Notification not found'}, status=404)
    
    
    
